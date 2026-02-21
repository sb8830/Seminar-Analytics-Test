"""
Enterprise Seminar Analytics Dashboard — ULTIMATE EDITION
==========================================================
Combines the best of all versions:
  ✅ OOP architecture (V3)
  ✅ 7-sheet styled Excel reports (V2)
  ✅ Dual KPI rows / 12 metrics (V3)
  ✅ Trend + Heatmap charts (V3)
  ✅ Trainer & Location report tabs (V2)
  ✅ Automated recommendations engine (V2)
  ✅ Attendance & Revenue range sliders (V3)
  ✅ Date range filter (V2+V3)
  ✅ CSV quick exports (V2+V3)
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from datetime import datetime
from typing import Dict, Any, Optional
import logging
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page Config ──
st.set_page_config(
    page_title="Seminar Analytics Pro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── CSS ──
st.markdown("""
<style>
    .block-container { padding-top: 1.2rem; padding-bottom: 2rem; max-width: 99%; }
    h1 { color: #1a1a2e; font-weight: 700; letter-spacing: -0.5px; }
    h2, h3 { color: #16213e; font-weight: 600; }
    div[data-testid="stMetric"] {
        background: #f8f9fa;
        padding: 14px;
        border-radius: 10px;
        border: 1px solid #e9ecef;
    }
    div[data-testid="stMetric"] label { color: #6b7280; font-size: 0.82rem; }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color: #1a1a2e; font-weight: 700; }
    .stTabs [data-baseweb="tab"] { padding: 10px 20px; border-radius: 8px; font-weight: 500; }
    .stDataFrame { border-radius: 10px; overflow: hidden; }
    .stFileUploader { background: #f8f9fa; padding: 15px; border-radius: 10px; border: 2px dashed #d1d5db; }
    .section-header {
        background: linear-gradient(90deg, #1a56db15, transparent);
        border-left: 4px solid #1a56db;
        padding: 8px 16px;
        border-radius: 0 8px 8px 0;
        margin: 12px 0 8px 0;
        font-weight: 600;
        color: #1a1a2e;
    }
</style>
""", unsafe_allow_html=True)

# ── Session State ──
for k, v in {'data_loaded': False, 'last_refresh': None}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ════════════════════════════════════════════════════
# DATA LAYER
# ════════════════════════════════════════════════════

class DataProcessor:
    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def load_attendee_data(file) -> pd.DataFrame:
        try:
            sheets = pd.read_excel(file, sheet_name=None, header=0)
            frames = []
            for name, df in sheets.items():
                df.columns = (df.columns.str.strip().str.lower()
                               .str.replace(' ', '_').str.replace(r'[^\w]', '_', regex=True))
                if any(c in df.columns for c in ['student_name', 'studentname', 'name']):
                    df['source_sheet'] = name
                    frames.append(df)
            if frames:
                result = pd.concat(frames, ignore_index=True)
                return DataProcessor._clean_attendee(result)
            return pd.DataFrame()
        except Exception as e:
            st.error(f"Failed to load attendee data: {e}")
            return pd.DataFrame()

    @staticmethod
    def _clean_attendee(df: pd.DataFrame) -> pd.DataFrame:
        aliases = {
            'studentname': 'student_name', 'name': 'student_name',
            'phone_no': 'phone', 'contact': 'phone', 'email_id': 'email',
            'service': 'service_name', 'course': 'service_name',
            'batch': 'batch_date', 'payment': 'payment_received',
            'amount_paid': 'payment_received', 'gst_amount': 'total_gst',
            'total': 'total_amount', 'due_amount': 'total_due', 'balance': 'total_due',
            'student_status': 'status', 'enrollment_status': 'status'
        }
        df = df.rename(columns={k: v for k, v in aliases.items() if k in df.columns})
        for col in ['payment_received', 'total_gst', 'total_amount', 'total_due', 'total_additional_charges']:
            if col in df.columns:
                df[col] = pd.to_numeric(
                    df[col].astype(str).str.replace(',', '').str.replace('₹', '').str.strip(),
                    errors='coerce').fillna(0)
        for col in ['student_name', 'email', 'service_name', 'status']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().replace('nan', '')
        return df

    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def load_seminar_data(file) -> pd.DataFrame:
        try:
            df = pd.read_excel(file, sheet_name=0, header=1)
            df.columns = df.columns.str.strip()
            return df
        except Exception as e:
            st.error(f"Failed to load seminar data: {e}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def parse_seminar(df: pd.DataFrame) -> pd.DataFrame:
        col_map = {}
        cols_lower = {c: c.strip().lower().replace('\n', ' ').replace('\r', ' ') for c in df.columns}

        patterns = {
            'sr_no':                  ['sr no', 'sr.no', 'serial', 's.no', 'srno'],
            'trainer':                ['trainer', 'faculty', 'speaker', 'mentor'],
            'location':               ['location', 'venue', 'city', 'branch'],
            'seminar_date':           ['seminar date', 'event date', 'date'],
            'targeted':               ['targeted', 'target', 'expected'],
            'total_attended':         ['total attended', 'total attendance', 'attended'],
            'actual_attended':        ['actual attended', 'actual attendance'],
            'targeted_to_attended_pct': ['targeted', 'attended', '%'],
            'total_seat_booked':      ['total seat booked', 'seats booked', 'booked seats'],
            'actual_expenses':        ['actual expense', 'expense', 'cost', 'actual cost'],
            'expected_revenue':       ['expected revenue', 'projected revenue'],
            'actual_revenue':         ['actual revenue', 'revenue', 'collection'],
            'total_revenue':          ['total revenue', 'total collection'],
            'surplus_deficit':        ['surplus', 'deficit', 'profit loss', 'p/l'],
            'er_to_ae':               ['er to ae', 'er/ae'],
            'ar_to_ae':               ['ar to ae', 'ar/ae'],
            'attended_to_booked_pct': ['attended to booked', 'conversion %'],
            'morning_total':          ['morning total', 'morning'],
            'evening_total':          ['evening total', 'evening'],
        }

        for target, keywords in patterns.items():
            for orig, low in cols_lower.items():
                if any(kw in low for kw in keywords) and target not in col_map:
                    col_map[target] = orig
                    break

        renamed = df.rename(columns={v: k for k, v in col_map.items() if v in df.columns})

        if 'sr_no' in renamed.columns:
            renamed = renamed[pd.to_numeric(renamed['sr_no'], errors='coerce').notna()]
            renamed['sr_no'] = pd.to_numeric(renamed['sr_no'], errors='coerce').astype(int)

        numeric_cols = [
            'targeted', 'total_attended', 'actual_attended', 'total_seat_booked',
            'actual_expenses', 'expected_revenue', 'actual_revenue', 'total_revenue',
            'surplus_deficit', 'er_to_ae', 'ar_to_ae', 'morning_total', 'evening_total'
        ]
        for col in numeric_cols:
            if col in renamed.columns:
                renamed[col] = pd.to_numeric(
                    renamed[col].astype(str).str.replace(',', '').str.replace('%', '').str.replace('₹', ''),
                    errors='coerce').fillna(0)

        if 'seminar_date' in renamed.columns:
            renamed['seminar_date'] = pd.to_datetime(renamed['seminar_date'], errors='coerce')

        return renamed


# ════════════════════════════════════════════════════
# ANALYTICS ENGINE
# ════════════════════════════════════════════════════

class AnalyticsEngine:
    @staticmethod
    def kpis(df: pd.DataFrame) -> Dict[str, Any]:
        if df.empty:
            return {}
        k = {}
        k['total_seminars'] = len(df)
        with_exp = df[df.get('actual_expenses', pd.Series(dtype=float)).gt(0)] if 'actual_expenses' in df.columns else df

        if 'total_attended' in df.columns:
            k['total_attended'] = int(df['total_attended'].sum())
            k['avg_attendance']  = round(df['total_attended'].mean(), 1)

        if 'actual_revenue' in df.columns:
            k['total_revenue'] = float(df['actual_revenue'].sum())
            k['avg_revenue']   = round(df['actual_revenue'].mean(), 0)

        if 'actual_expenses' in df.columns:
            k['total_expenses'] = float(df['actual_expenses'].sum())
            k['avg_expense']    = round(df['actual_expenses'].mean(), 0)

        if 'surplus_deficit' in df.columns:
            k['total_profit']        = float(df['surplus_deficit'].sum())
            k['profitable_seminars'] = int((with_exp['surplus_deficit'] > 0).sum()) if not with_exp.empty else 0
            k['loss_seminars']       = int((with_exp['surplus_deficit'] < 0).sum()) if not with_exp.empty else 0
            k['profit_margin']       = round(k['total_profit'] / k['total_revenue'] * 100, 1) if k.get('total_revenue', 0) > 0 else 0

        if 'total_seat_booked' in df.columns and 'total_attended' in df.columns:
            att = df['total_attended'].sum()
            bkd = df['total_seat_booked'].sum()
            k['total_booked']    = int(bkd)
            k['conversion_rate'] = round(bkd / att * 100 if att > 0 else 0, 1)

        if 'targeted' in df.columns and 'total_attended' in df.columns:
            k['target_achievement'] = round(
                df['total_attended'].sum() / df['targeted'].sum() * 100
                if df['targeted'].sum() > 0 else 0, 1)

        if k.get('total_expenses', 0) > 0 and k.get('total_revenue', 0) > 0:
            k['expense_ratio'] = round(k['total_expenses'] / k['total_revenue'] * 100, 1)
        else:
            k['expense_ratio'] = 0

        return k

    @staticmethod
    def location_summary(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or 'location' not in df.columns:
            return pd.DataFrame()
        agg_dict = {}
        for col, func in [('sr_no', 'count'), ('total_attended', 'sum'), ('total_seat_booked', 'sum'),
                           ('actual_revenue', 'sum'), ('actual_expenses', 'sum'), ('surplus_deficit', 'sum')]:
            if col in df.columns:
                agg_dict[col] = func
        if not agg_dict:
            return pd.DataFrame()
        g = df.groupby('location').agg(agg_dict).reset_index()
        g = g.rename(columns={'sr_no': 'seminars'})
        if 'actual_revenue' in g.columns and 'surplus_deficit' in g.columns:
            g['profit_margin_pct'] = (g['surplus_deficit'] / g['actual_revenue'].replace(0, np.nan) * 100).round(1).fillna(0)
        if 'total_seat_booked' in g.columns and 'total_attended' in g.columns:
            g['conversion_pct'] = (g['total_seat_booked'] / g['total_attended'].replace(0, np.nan) * 100).round(1).fillna(0)
        return g.sort_values('actual_revenue', ascending=False)

    @staticmethod
    def trainer_summary(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty or 'trainer' not in df.columns:
            return pd.DataFrame()
        rows = []
        for _, row in df.iterrows():
            trainers = [t.strip().split('\n')[0].strip()
                        for t in str(row.get('trainer', '')).split(',') if t.strip()]
            for t in trainers:
                if t:
                    rows.append({
                        'trainer': t,
                        'seminars': 1,
                        'total_attended':    row.get('total_attended', 0),
                        'total_seat_booked': row.get('total_seat_booked', 0),
                        'actual_revenue':    row.get('actual_revenue', 0),
                        'actual_expenses':   row.get('actual_expenses', 0),
                        'surplus_deficit':   row.get('surplus_deficit', 0),
                    })
        if not rows:
            return pd.DataFrame()
        tdf = pd.DataFrame(rows).groupby('trainer').sum().reset_index()
        tdf['avg_rev_per_seminar'] = (tdf['actual_revenue'] / tdf['seminars'].replace(0, np.nan)).round(0).fillna(0)
        tdf['conversion_pct'] = (tdf['total_seat_booked'] / tdf['total_attended'].replace(0, np.nan) * 100).round(1).fillna(0)
        return tdf.sort_values('actual_revenue', ascending=False)

    @staticmethod
    def course_summary(attendee_df: pd.DataFrame) -> pd.DataFrame:
        if attendee_df.empty or 'service_name' not in attendee_df.columns:
            return pd.DataFrame()
        agg = {}
        if 'student_name' in attendee_df.columns:
            agg['student_name'] = 'count'
        if 'payment_received' in attendee_df.columns:
            agg['payment_received'] = 'sum'
        if 'total_gst' in attendee_df.columns:
            agg['total_gst'] = 'sum'
        if 'total_due' in attendee_df.columns:
            agg['total_due'] = 'sum'
        if not agg:
            return pd.DataFrame()
        g = attendee_df.groupby('service_name').agg(agg).reset_index()
        g = g.rename(columns={'student_name': 'enrollments', 'payment_received': 'revenue'})
        return g.sort_values('revenue', ascending=False)


# ════════════════════════════════════════════════════
# CHARTS
# ════════════════════════════════════════════════════

COLORS = {
    'blue': '#1a56db', 'green': '#10b981', 'red': '#ef4444',
    'amber': '#f59e0b', 'purple': '#6366f1', 'gray': '#9ca3af',
    'teal': '#14b8a6', 'pink': '#ec4899'
}
CHART_LAYOUT = dict(
    plot_bgcolor='rgba(0,0,0,0)',
    paper_bgcolor='rgba(0,0,0,0)',
    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
    margin=dict(l=20, r=20, t=50, b=80),
    height=420
)

def fmt_hover(v):
    return f"₹{v:,.0f}"

def chart_rev_expense(df):
    data = df[df.get('actual_expenses', pd.Series(0, index=df.index)).gt(0)] if 'actual_expenses' in df.columns else df
    if data.empty or not all(c in data.columns for c in ['location', 'actual_revenue', 'actual_expenses']):
        return None
    fig = go.Figure([
        go.Bar(name='Revenue', x=data['location'], y=data['actual_revenue'],
               marker_color=COLORS['blue'], hovertemplate='₹%{y:,.0f}<extra>Revenue</extra>'),
        go.Bar(name='Expense', x=data['location'], y=data['actual_expenses'],
               marker_color=COLORS['amber'], hovertemplate='₹%{y:,.0f}<extra>Expense</extra>')
    ])
    fig.update_layout(barmode='group', xaxis_tickangle=-40, title='Revenue vs Expense by Location', **CHART_LAYOUT)
    return fig

def chart_surplus(df):
    if 'surplus_deficit' not in df.columns:
        return None
    data = df[df.get('actual_expenses', pd.Series(0, index=df.index)).gt(0)][['location', 'surplus_deficit']].copy() if 'actual_expenses' in df.columns else df[['location', 'surplus_deficit']].copy()
    if data.empty:
        return None
    colors = [COLORS['green'] if v >= 0 else COLORS['red'] for v in data['surplus_deficit']]
    fig = px.bar(data, x='location', y='surplus_deficit')
    fig.update_traces(marker_color=colors, hovertemplate='₹%{y:,.0f}<extra>Surplus/Deficit</extra>')
    fig.add_hline(y=0, line_dash='dash', line_color='gray', line_width=1)
    fig.update_layout(xaxis_tickangle=-40, title='Surplus / Deficit by Location', **CHART_LAYOUT)
    return fig

def chart_funnel(df):
    needed = ['targeted', 'total_attended', 'total_seat_booked', 'location']
    if not all(c in df.columns for c in needed) or df.empty:
        return None
    fig = go.Figure([
        go.Bar(name='Targeted', x=df['location'], y=df['targeted'], marker_color=COLORS['gray'], opacity=0.5),
        go.Bar(name='Attended', x=df['location'], y=df['total_attended'], marker_color=COLORS['blue']),
        go.Bar(name='Booked',   x=df['location'], y=df['total_seat_booked'], marker_color=COLORS['green'])
    ])
    fig.update_layout(barmode='group', xaxis_tickangle=-40, title='Attendance Funnel by Location', **CHART_LAYOUT)
    return fig

def chart_trend(df):
    if 'seminar_date' not in df.columns or df.empty:
        return None
    ds = df.sort_values('seminar_date')
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    if 'total_attended' in ds.columns:
        fig.add_trace(go.Bar(x=ds['seminar_date'], y=ds['total_attended'], name='Attendance',
                             marker_color=COLORS['blue'], opacity=0.7), secondary_y=False)
    if 'surplus_deficit' in ds.columns:
        fig.add_trace(go.Scatter(x=ds['seminar_date'], y=ds['surplus_deficit'], name='Profit/Loss',
                                 mode='lines+markers', line=dict(color=COLORS['green'], width=2.5),
                                 marker=dict(size=7)), secondary_y=True)
    if 'actual_revenue' in ds.columns:
        fig.add_trace(go.Scatter(x=ds['seminar_date'], y=ds['actual_revenue'], name='Revenue',
                                 mode='lines', line=dict(color=COLORS['amber'], width=2, dash='dot')), secondary_y=True)
    fig.update_layout(title='Performance Trend Over Time', **CHART_LAYOUT)
    fig.update_yaxes(title_text='Attendance', secondary_y=False)
    fig.update_yaxes(title_text='₹ Revenue / Profit', secondary_y=True)
    return fig

def chart_heatmap(df):
    if df.empty or 'location' not in df.columns:
        return None
    cols = [c for c in ['total_attended', 'actual_revenue', 'actual_expenses', 'surplus_deficit', 'total_seat_booked'] if c in df.columns]
    if not cols:
        return None
    g = df.groupby('location')[cols].sum()
    norm = g.apply(lambda col: col / col.abs().max() * 100 if col.abs().max() > 0 else col, axis=0)
    labels = {'total_attended': 'Attendance', 'actual_revenue': 'Revenue', 'actual_expenses': 'Expenses',
              'surplus_deficit': 'Profit', 'total_seat_booked': 'Booked'}
    fig = px.imshow(
        norm.T,
        labels=dict(x='Location', y='Metric', color='Score (%)'),
        x=norm.index.tolist(),
        y=[labels.get(c, c) for c in cols],
        color_continuous_scale='RdYlGn',
        title='Location Performance Heatmap'
    )
    fig.update_layout(height=380)
    return fig

def chart_status_pie(attendee_df):
    if 'status' not in attendee_df.columns or attendee_df.empty:
        return None
    sc = attendee_df['status'].value_counts()
    if sc.empty:
        return None
    fig = px.pie(values=sc.values, names=sc.index, hole=0.45, title='Student Status Distribution',
                 color_discrete_sequence=[COLORS['green'], COLORS['red'], COLORS['amber'],
                                          COLORS['purple'], COLORS['teal']])
    fig.update_layout(height=400)
    return fig

def chart_course_revenue(attendee_df):
    if 'service_name' not in attendee_df.columns or 'payment_received' not in attendee_df.columns:
        return None
    cr = attendee_df.groupby('service_name')['payment_received'].sum().sort_values(ascending=False).head(12)
    if cr.empty:
        return None
    fig = px.bar(x=cr.values, y=cr.index, orientation='h', title='Top 12 Courses by Revenue',
                 color=cr.values, color_continuous_scale='Blues')
    fig.update_layout(yaxis_title='', xaxis_title='Revenue (₹)',
                      coloraxis_showscale=False, **CHART_LAYOUT)
    return fig

def chart_trainer_performance(trainer_df):
    if trainer_df.empty:
        return None
    top = trainer_df.head(10)
    fig = make_subplots(rows=1, cols=2, subplot_titles=('Revenue by Trainer', 'Seminars Conducted'))
    fig.add_trace(go.Bar(x=top['trainer'], y=top['actual_revenue'], name='Revenue',
                         marker_color=COLORS['blue']), row=1, col=1)
    fig.add_trace(go.Bar(x=top['trainer'], y=top['seminars'], name='Seminars',
                         marker_color=COLORS['purple']), row=1, col=2)
    fig.update_layout(showlegend=False, height=400,
                      plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=20, r=20, t=60, b=80))
    fig.update_xaxes(tickangle=-30)
    return fig

def chart_location_bubble(loc_df):
    if loc_df.empty or not all(c in loc_df.columns for c in ['total_attended', 'actual_revenue', 'surplus_deficit']):
        return None
    fig = px.scatter(loc_df, x='total_attended', y='actual_revenue',
                     size=loc_df['surplus_deficit'].clip(lower=1),
                     color='profit_margin_pct' if 'profit_margin_pct' in loc_df.columns else 'actual_revenue',
                     hover_name='location', title='Attendance vs Revenue (bubble = profit size)',
                     color_continuous_scale='RdYlGn', size_max=50)
    fig.update_layout(height=400, plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
    return fig


# ════════════════════════════════════════════════════
# EXCEL REPORT GENERATOR
# ════════════════════════════════════════════════════

def generate_excel_report(filtered_df, attendee_df, trainer_df, location_df, course_df, kpi_dict):
    wb = openpyxl.Workbook()

    # Shared styles
    def hdr_style(cell, bg='1A56DB', fg='FFFFFF', size=10, bold=True):
        cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()

    def _border(color='CCCCCC'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def data_style(cell, alt=False, bold=False):
        cell.font = Font(name='Arial', size=9, bold=bold)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = _border()
        if alt:
            cell.fill = PatternFill('solid', start_color='F0F4FF')

    def pos_neg_fill(cell, val):
        if val > 0:
            cell.fill = PatternFill('solid', start_color='D1FAE5')
        elif val < 0:
            cell.fill = PatternFill('solid', start_color='FEE2E2')

    def add_title(ws, text, sub=''):
        ws.merge_cells('A1:J1')
        ws['A1'] = text
        ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        if sub:
            ws.merge_cells('A2:J2')
            ws['A2'] = sub
            ws['A2'].font = Font(name='Arial', italic=True, size=9, color='6B7280')
            ws['A2'].alignment = Alignment(horizontal='center')

    def write_table(ws, headers, rows, start_row, col_count, highlight_col=None):
        for ci, h in enumerate(headers, 1):
            hdr_style(ws.cell(row=start_row, column=ci), size=9)
            ws.cell(row=start_row, column=ci).value = h
        for ri, row_vals in enumerate(rows, start_row + 1):
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci)
                c.value = val if not (isinstance(val, float) and np.isnan(val)) else ''
                data_style(c, alt=(ri % 2 == 0))
            if highlight_col and highlight_col <= len(row_vals):
                pos_neg_fill(ws.cell(row=ri, column=highlight_col), row_vals[highlight_col - 1] or 0)

    ts = datetime.now().strftime('%d %b %Y %I:%M %p')

    # ── Sheet 1: Executive Summary ──
    ws1 = wb.active
    ws1.title = 'Executive Summary'
    add_title(ws1, '📊 EXECUTIVE SUMMARY', f'Generated: {ts}')

    kpi_rows = [
        ('Total Seminars',        kpi_dict.get('total_seminars', 0),          ''),
        ('Total Attendees',       kpi_dict.get('total_attended', 0),           ''),
        ('Total Revenue',         f"₹{kpi_dict.get('total_revenue', 0):,.0f}", ''),
        ('Total Expenses',        f"₹{kpi_dict.get('total_expenses', 0):,.0f}", ''),
        ('Net Profit / Deficit',  f"₹{kpi_dict.get('total_profit', 0):,.0f}",  '▲ Profit' if kpi_dict.get('total_profit', 0) >= 0 else '▼ Loss'),
        ('Profit Margin',         f"{kpi_dict.get('profit_margin', 0):.1f}%",  ''),
        ('Avg Revenue / Seminar', f"₹{kpi_dict.get('avg_revenue', 0):,.0f}",  ''),
        ('Conversion Rate',       f"{kpi_dict.get('conversion_rate', 0):.1f}%", ''),
        ('Target Achievement',    f"{kpi_dict.get('target_achievement', 0):.1f}%", ''),
        ('Profitable Seminars',   f"{kpi_dict.get('profitable_seminars', 0)} / {kpi_dict.get('total_seminars', 0)}", ''),
        ('Loss-making Seminars',  kpi_dict.get('loss_seminars', 0),            ''),
        ('Expense Ratio',         f"{kpi_dict.get('expense_ratio', 0):.1f}%",  'Target: <50%'),
    ]
    write_table(ws1, ['KPI Metric', 'Value', 'Note'], kpi_rows, start_row=4, col_count=3)
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 18

    # Top / Bottom 5
    if 'surplus_deficit' in filtered_df.columns and len(filtered_df) > 0:
        wexp = filtered_df[filtered_df.get('actual_expenses', pd.Series(0)).gt(0)] if 'actual_expenses' in filtered_df.columns else filtered_df
        if not wexp.empty:
            r = 18
            ws1.cell(row=r, column=1).value = '🏆 Top 5 Profitable Locations'
            ws1.cell(row=r, column=1).font = Font(name='Arial', bold=True, color='1A56DB', size=10)
            ws1.merge_cells(f'A{r}:C{r}')
            top5 = wexp.nlargest(5, 'surplus_deficit')
            write_table(ws1, ['Location', 'Surplus (₹)', 'Status'],
                        [(row.get('location', ''), int(row['surplus_deficit']), '✅ Profit') for _, row in top5.iterrows()],
                        start_row=r + 1, col_count=3, highlight_col=2)
            r2 = r + 8
            ws1.cell(row=r2, column=1).value = '⚠️ Bottom 5 Locations'
            ws1.cell(row=r2, column=1).font = Font(name='Arial', bold=True, color='EF4444', size=10)
            ws1.merge_cells(f'A{r2}:C{r2}')
            bot5 = wexp.nsmallest(5, 'surplus_deficit')
            write_table(ws1, ['Location', 'Deficit (₹)', 'Status'],
                        [(row.get('location', ''), int(row['surplus_deficit']), '❌ Loss' if row['surplus_deficit'] < 0 else '~Break-even') for _, row in bot5.iterrows()],
                        start_row=r2 + 1, col_count=3, highlight_col=2)

    # ── Sheet 2: Financial Report ──
    ws2 = wb.create_sheet('Financial Report')
    add_title(ws2, '💰 FINANCIAL PERFORMANCE REPORT', ts)
    fcols = [c for c in ['sr_no', 'location', 'trainer', 'seminar_date', 'actual_expenses',
                          'expected_revenue', 'actual_revenue', 'surplus_deficit', 'er_to_ae', 'ar_to_ae']
             if c in filtered_df.columns]
    fhdrs = {'sr_no': 'Sr', 'location': 'Location', 'trainer': 'Trainer', 'seminar_date': 'Date',
             'actual_expenses': 'Expenses (₹)', 'expected_revenue': 'Exp Rev (₹)',
             'actual_revenue': 'Act Rev (₹)', 'surplus_deficit': 'Surplus/Deficit (₹)',
             'er_to_ae': 'ER:AE', 'ar_to_ae': 'AR:AE'}
    for ci, col in enumerate(fcols, 1):
        c = ws2.cell(row=4, column=ci)
        c.value = fhdrs.get(col, col.replace('_', ' ').title())
        hdr_style(c, size=9)
    sd_idx = fcols.index('surplus_deficit') + 1 if 'surplus_deficit' in fcols else None
    for ri, (_, row) in enumerate(filtered_df[fcols].iterrows(), 5):
        for ci, col in enumerate(fcols, 1):
            val = row[col]
            c = ws2.cell(row=ri, column=ci)
            c.value = '' if pd.isna(val) else val
            data_style(c, alt=(ri % 2 == 0))
        if sd_idx:
            pos_neg_fill(ws2.cell(row=ri, column=sd_idx), row.get('surplus_deficit', 0))
    tr = ri + 2
    ws2.cell(row=tr, column=1).value = 'TOTAL'
    ws2.cell(row=tr, column=1).font = Font(name='Arial', bold=True, size=10)
    for ci, col in enumerate(fcols, 1):
        if col in ['actual_expenses', 'expected_revenue', 'actual_revenue', 'surplus_deficit']:
            c = ws2.cell(row=tr, column=ci)
            c.value = f'=SUM({get_column_letter(ci)}5:{get_column_letter(ci)}{ri})'
            c.font = Font(name='Arial', bold=True, size=10)
            c.fill = PatternFill('solid', start_color='DBEAFE')
    for ci in range(1, len(fcols) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 16
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24

    # ── Sheet 3: Attendance & Conversion ──
    ws3 = wb.create_sheet('Attendance & Conversion')
    add_title(ws3, '🎯 ATTENDANCE & CONVERSION REPORT', ts)
    acols = [c for c in ['sr_no', 'location', 'seminar_date', 'targeted', 'total_attended',
                          'actual_attended', 'total_seat_booked', 'morning_total', 'evening_total']
             if c in filtered_df.columns]
    ahdrs = {'sr_no': 'Sr', 'location': 'Location', 'seminar_date': 'Date', 'targeted': 'Targeted',
             'total_attended': 'Attended', 'actual_attended': 'Actual Att.', 'total_seat_booked': 'Booked',
             'morning_total': 'Morning', 'evening_total': 'Evening'}
    for ci, col in enumerate(acols, 1):
        c = ws3.cell(row=4, column=ci)
        c.value = ahdrs.get(col, col.replace('_', ' ').title())
        hdr_style(c, size=9)
    extra = len(acols) + 1
    c = ws3.cell(row=4, column=extra)
    c.value = 'Conversion %'
    hdr_style(c, size=9)
    for ri, (_, row) in enumerate(filtered_df[acols].iterrows(), 5):
        for ci, col in enumerate(acols, 1):
            val = row[col]
            c = ws3.cell(row=ri, column=ci)
            c.value = '' if pd.isna(val) else val
            data_style(c, alt=(ri % 2 == 0))
        if 'total_seat_booked' in acols and 'total_attended' in acols:
            bc = get_column_letter(acols.index('total_seat_booked') + 1)
            ac = get_column_letter(acols.index('total_attended') + 1)
            ws3.cell(row=ri, column=extra).value = f'=IF({ac}{ri}>0,{bc}{ri}/{ac}{ri}*100,0)'
    for ci in range(1, extra + 1):
        ws3.column_dimensions[get_column_letter(ci)].width = 16
    ws3.column_dimensions['B'].width = 24

    # ── Sheet 4: Trainer Performance ──
    ws4 = wb.create_sheet('Trainer Performance')
    add_title(ws4, '👨‍🏫 TRAINER PERFORMANCE REPORT', ts)
    if not trainer_df.empty:
        th = ['Trainer', 'Seminars', 'Attended', 'Booked', 'Revenue (₹)', 'Expenses (₹)', 'Surplus (₹)', 'Avg Rev/Sem', 'Conversion %']
        for ci, h in enumerate(th, 1):
            c = ws4.cell(row=4, column=ci)
            c.value = h
            hdr_style(c, size=9)
        for ri, (_, row) in enumerate(trainer_df.iterrows(), 5):
            vals = [row['trainer'], int(row['seminars']), int(row['total_attended']),
                    int(row['total_seat_booked']), round(row['actual_revenue']),
                    round(row['actual_expenses']), round(row['surplus_deficit']),
                    f'=IF(B{ri}>0,E{ri}/B{ri},0)', f'=IF(C{ri}>0,D{ri}/C{ri}*100,0)']
            for ci, val in enumerate(vals, 1):
                c = ws4.cell(row=ri, column=ci)
                c.value = val
                data_style(c, alt=(ri % 2 == 0))
            pos_neg_fill(ws4.cell(row=ri, column=7), row['surplus_deficit'])
        for ci, w in enumerate([30, 12, 14, 12, 18, 16, 18, 18, 16], 1):
            ws4.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 5: Location Summary ──
    ws5 = wb.create_sheet('Location Summary')
    add_title(ws5, '📍 LOCATION-WISE SUMMARY', ts)
    if not location_df.empty:
        lh = ['Location', 'Seminars', 'Attended', 'Booked', 'Revenue (₹)', 'Expenses (₹)',
              'Surplus (₹)', 'Profit Margin %', 'Conversion %']
        for ci, h in enumerate(lh, 1):
            c = ws5.cell(row=4, column=ci)
            c.value = h
            hdr_style(c, size=9)
        for ri, (_, row) in enumerate(location_df.iterrows(), 5):
            vals = [row['location'],
                    int(row.get('seminars', 0)), int(row.get('total_attended', 0)),
                    int(row.get('total_seat_booked', 0)), round(row.get('actual_revenue', 0)),
                    round(row.get('actual_expenses', 0)), round(row.get('surplus_deficit', 0)),
                    round(row.get('profit_margin_pct', 0), 1), round(row.get('conversion_pct', 0), 1)]
            for ci, val in enumerate(vals, 1):
                c = ws5.cell(row=ri, column=ci)
                c.value = val
                data_style(c, alt=(ri % 2 == 0))
            pos_neg_fill(ws5.cell(row=ri, column=7), row.get('surplus_deficit', 0))
            pos_neg_fill(ws5.cell(row=ri, column=8), row.get('profit_margin_pct', 0))
        for ci in range(1, 10):
            ws5.column_dimensions[get_column_letter(ci)].width = 20
        ws5.column_dimensions['A'].width = 28

    # ── Sheet 6: Attendee Details ──
    ws6 = wb.create_sheet('Attendee Details')
    add_title(ws6, '👤 FULL ATTENDEE DETAILS', ts)
    att_cols = [c for c in ['student_name', 'phone', 'email', 'service_name', 'batch_date',
                             'payment_received', 'total_gst', 'total_amount', 'total_due', 'status']
                if c in attendee_df.columns]
    att_hdrs = {'student_name': 'Student', 'phone': 'Phone', 'email': 'Email', 'service_name': 'Course',
                'batch_date': 'Batch Date', 'payment_received': 'Payment (₹)', 'total_gst': 'GST (₹)',
                'total_amount': 'Total (₹)', 'total_due': 'Due (₹)', 'status': 'Status'}
    for ci, col in enumerate(att_cols, 1):
        c = ws6.cell(row=4, column=ci)
        c.value = att_hdrs.get(col, col.replace('_', ' ').title())
        hdr_style(c, size=9)
    for ri, (_, row) in enumerate(attendee_df[att_cols].iterrows(), 5):
        for ci, col in enumerate(att_cols, 1):
            val = row[col]
            c = ws6.cell(row=ri, column=ci)
            c.value = '' if pd.isna(val) else val
            c.font = Font(name='Arial', size=9)
            c.border = _border()
            if ri % 2 == 0:
                c.fill = PatternFill('solid', start_color='F8FAFF')
    for ci in range(1, len(att_cols) + 1):
        ws6.column_dimensions[get_column_letter(ci)].width = 20
    ws6.column_dimensions['A'].width = 28
    if 'C' in ws6.column_dimensions:
        ws6.column_dimensions['C'].width = 32

    # ── Sheet 7: Course Revenue ──
    ws7 = wb.create_sheet('Course Revenue')
    add_title(ws7, '📚 COURSE-WISE REVENUE BREAKDOWN', ts)
    if not course_df.empty:
        ch = ['Course / Service', 'Enrollments', 'Revenue (₹)', 'GST (₹)', 'Total Due (₹)', 'Avg Rev/Student']
        for ci, h in enumerate(ch, 1):
            c = ws7.cell(row=4, column=ci)
            c.value = h
            hdr_style(c, size=9)
        ri = 5
        for _, row in course_df.iterrows():
            ws7.cell(row=ri, column=1).value = row['service_name']
            ws7.cell(row=ri, column=2).value = int(row.get('enrollments', 0))
            ws7.cell(row=ri, column=3).value = round(row.get('revenue', 0))
            ws7.cell(row=ri, column=4).value = round(row.get('total_gst', 0))
            ws7.cell(row=ri, column=5).value = round(row.get('total_due', 0))
            ws7.cell(row=ri, column=6).value = f'=IF(B{ri}>0,C{ri}/B{ri},0)'
            for ci in range(1, 7):
                data_style(ws7.cell(row=ri, column=ci), alt=(ri % 2 == 0))
            ri += 1
        # Totals
        ws7.cell(row=ri + 1, column=1).value = 'GRAND TOTAL'
        ws7.cell(row=ri + 1, column=1).font = Font(name='Arial', bold=True, size=10)
        for ci in [2, 3, 4, 5]:
            c = ws7.cell(row=ri + 1, column=ci)
            c.value = f'=SUM({get_column_letter(ci)}5:{get_column_letter(ci)}{ri})'
            c.font = Font(name='Arial', bold=True)
            c.fill = PatternFill('solid', start_color='DBEAFE')
            c.border = _border()
        for ci, w in enumerate([36, 14, 18, 14, 16, 22], 1):
            ws7.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════

def fmt_currency(v):
    v = float(v) if v else 0
    if abs(v) >= 1e7:   return f"₹{v/1e7:.2f}Cr"
    if abs(v) >= 1e5:   return f"₹{v/1e5:.1f}L"
    if abs(v) >= 1e3:   return f"₹{v/1e3:.1f}K"
    return f"₹{v:,.0f}"

def fmt_num(v):
    v = int(v) if v else 0
    if abs(v) >= 1e6: return f"{v/1e6:.1f}M"
    if abs(v) >= 1e3: return f"{v/1e3:.1f}K"
    return f"{v:,}"

def show_chart(fig, key):
    if fig:
        st.plotly_chart(fig, use_container_width=True, key=key)
    else:
        st.info("Not enough data to render this chart.")


# ════════════════════════════════════════════════════
# MAIN APP
# ════════════════════════════════════════════════════

def main():
    st.title("📊 Seminar Analytics Pro")
    st.caption("Offline Seminar Performance • 2025-26 | Enterprise Analytics")

    # ── Upload ──
    st.markdown("### 📁 Data Import")
    c1, c2, c3 = st.columns([2, 2, 1])
    with c1:
        file1 = st.file_uploader("Upload **Attendee Details** (.xlsx)", type=["xlsx", "xls"], key="f1")
    with c2:
        file2 = st.file_uploader("Upload **Seminar Report** (.xlsx)", type=["xlsx", "xls"], key="f2")
    with c3:
        st.write("")
        st.write("")
        if st.button("🔄 Refresh Cache", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    if not file1 or not file2:
        st.info("👆 Please upload both Excel files to begin.")
        st.stop()

    # ── Load Data ──
    with st.spinner("Loading & processing data…"):
        attendee_df  = DataProcessor.load_attendee_data(file1)
        seminar_raw  = DataProcessor.load_seminar_data(file2)
        seminar_df   = DataProcessor.parse_seminar(seminar_raw)

    if seminar_df.empty:
        st.error("❌ No valid seminar data found. Please check your file format.")
        st.stop()

    st.session_state['data_loaded']  = True
    st.session_state['last_refresh'] = datetime.now().strftime('%d %b %Y %I:%M %p')

    # ── Sidebar Filters ──
    st.sidebar.header("🔍 Filters")
    st.sidebar.markdown("---")

    locs = sorted(seminar_df['location'].dropna().unique()) if 'location' in seminar_df.columns else []
    sel_locs = st.sidebar.multiselect("📍 Location", locs, default=locs)

    all_trainers = set()
    if 'trainer' in seminar_df.columns:
        for t in seminar_df['trainer'].dropna():
            for name in str(t).split(','):
                n = name.strip().split('\n')[0].strip()
                if n: all_trainers.add(n)
    sel_trainers = st.sidebar.multiselect("👨‍🏫 Trainer", sorted(all_trainers), default=[])

    if 'seminar_date' in seminar_df.columns:
        vd = seminar_df['seminar_date'].dropna()
        if not vd.empty:
            date_range = st.sidebar.date_input("📅 Date Range",
                                               value=(vd.min().date(), vd.max().date()),
                                               min_value=vd.min().date(), max_value=vd.max().date())
        else:
            date_range = None
    else:
        date_range = None

    profit_filter = st.sidebar.radio("💰 Profitability", ["All", "Profitable", "Loss-making"], horizontal=True)

    if 'total_attended' in seminar_df.columns:
        att_min, att_max = int(seminar_df['total_attended'].min()), int(seminar_df['total_attended'].max())
        if att_min < att_max:
            att_range = st.sidebar.slider("👥 Attendance Range", att_min, att_max, (att_min, att_max))
        else:
            att_range = (att_min, att_max)
    else:
        att_range = None

    if 'actual_revenue' in seminar_df.columns:
        rev_min, rev_max = int(seminar_df['actual_revenue'].min()), int(seminar_df['actual_revenue'].max())
        if rev_min < rev_max:
            rev_range = st.sidebar.slider("💰 Revenue Range (₹)", rev_min, rev_max, (rev_min, rev_max))
        else:
            rev_range = (rev_min, rev_max)
    else:
        rev_range = None

    st.sidebar.markdown("---")
    st.sidebar.caption(f"🕐 Last refreshed: {st.session_state.get('last_refresh', 'N/A')}")

    # ── Apply Filters ──
    filtered = seminar_df.copy()
    if sel_locs:
        filtered = filtered[filtered['location'].isin(sel_locs)]
    if sel_trainers:
        def has_trainer(ts):
            return any(n.strip().split('\n')[0].strip() in sel_trainers for n in str(ts).split(','))
        filtered = filtered[filtered['trainer'].apply(has_trainer)]
    if date_range and len(date_range) == 2 and 'seminar_date' in filtered.columns:
        filtered = filtered[
            (filtered['seminar_date'] >= pd.to_datetime(date_range[0])) &
            (filtered['seminar_date'] <= pd.to_datetime(date_range[1]))
        ]
    if profit_filter == "Profitable" and 'surplus_deficit' in filtered.columns:
        filtered = filtered[filtered['surplus_deficit'] > 0]
    elif profit_filter == "Loss-making" and 'surplus_deficit' in filtered.columns:
        filtered = filtered[filtered['surplus_deficit'] < 0]
    if att_range and 'total_attended' in filtered.columns:
        filtered = filtered[filtered['total_attended'].between(*att_range)]
    if rev_range and 'actual_revenue' in filtered.columns:
        filtered = filtered[filtered['actual_revenue'].between(*rev_range)]

    # ── Pre-compute Summaries ──
    kpi           = AnalyticsEngine.kpis(filtered)
    loc_summary   = AnalyticsEngine.location_summary(filtered)
    train_summary = AnalyticsEngine.trainer_summary(filtered)
    course_sum    = AnalyticsEngine.course_summary(attendee_df)
    with_exp      = filtered[filtered['actual_expenses'] > 0] if 'actual_expenses' in filtered.columns else filtered

    # ════════════════════════════════════════════════
    # KPI SECTION
    # ════════════════════════════════════════════════
    st.markdown("---")
    st.markdown('<div class="section-header">📈 Key Performance Indicators</div>', unsafe_allow_html=True)

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("📋 Seminars",       kpi.get('total_seminars', 0))
    k2.metric("👥 Attendees",      fmt_num(kpi.get('total_attended', 0)))
    k3.metric("💰 Revenue",        fmt_currency(kpi.get('total_revenue', 0)))
    k4.metric("📤 Expenses",       fmt_currency(kpi.get('total_expenses', 0)))
    k5.metric("📈 Net Profit",     fmt_currency(kpi.get('total_profit', 0)),
              delta="Profit" if kpi.get('total_profit', 0) >= 0 else "Loss")
    k6.metric("✅ Profitable",     f"{kpi.get('profitable_seminars', 0)}/{kpi.get('total_seminars', 0)}")

    s1, s2, s3, s4, s5, s6 = st.columns(6)
    s1.metric("📊 Avg Attendance",   kpi.get('avg_attendance', 0))
    s2.metric("💵 Avg Revenue",      fmt_currency(kpi.get('avg_revenue', 0)))
    s3.metric("📉 Profit Margin",    f"{kpi.get('profit_margin', 0):.1f}%")
    s4.metric("🎯 Conversion Rate",  f"{kpi.get('conversion_rate', 0):.1f}%")
    s5.metric("🎖️ Target Achieved",  f"{kpi.get('target_achievement', 0):.1f}%")
    s6.metric("📉 Loss-making",      kpi.get('loss_seminars', 0))

    # ════════════════════════════════════════════════
    # MAIN TABS
    # ════════════════════════════════════════════════
    st.markdown("---")
    tab_charts, tab_loc, tab_trainer, tab_insights, tab_tables, tab_reports = st.tabs([
        "📊 Charts & Analytics",
        "📍 Location Report",
        "👨‍🏫 Trainer Report",
        "💡 Insights",
        "📋 Data Tables",
        "📥 Download Reports"
    ])

    # ────────────────────────────────────────────────
    # TAB 1 — CHARTS
    # ────────────────────────────────────────────────
    with tab_charts:
        ct1, ct2, ct3, ct4, ct5, ct6, ct7 = st.tabs([
            "💰 Revenue vs Expense", "📈 Surplus/Deficit", "🎯 Attendance Funnel",
            "📅 Trend", "🗺️ Heatmap", "👥 Student Status", "📚 Course Revenue"
        ])
        with ct1: show_chart(chart_rev_expense(filtered), 'rev_exp')
        with ct2: show_chart(chart_surplus(filtered), 'surplus')
        with ct3: show_chart(chart_funnel(filtered), 'funnel')
        with ct4: show_chart(chart_trend(filtered), 'trend')
        with ct5: show_chart(chart_heatmap(filtered), 'heatmap')
        with ct6: show_chart(chart_status_pie(attendee_df), 'status_pie')
        with ct7: show_chart(chart_course_revenue(attendee_df), 'course_rev')

    # ────────────────────────────────────────────────
    # TAB 2 — LOCATION
    # ────────────────────────────────────────────────
    with tab_loc:
        st.markdown('<div class="section-header">📍 Location-wise Performance</div>', unsafe_allow_html=True)
        if not loc_summary.empty:
            lc1, lc2 = st.columns(2)
            with lc1:
                show_chart(chart_rev_expense(
                    loc_summary.rename(columns={'actual_revenue': 'actual_revenue',
                                                'actual_expenses': 'actual_expenses',
                                                'location': 'location'})), 'loc_rev')
            with lc2:
                show_chart(chart_location_bubble(loc_summary), 'loc_bubble')

            st.dataframe(
                loc_summary.style.format({
                    'actual_revenue':    '₹{:,.0f}', 'actual_expenses': '₹{:,.0f}',
                    'surplus_deficit':   '₹{:,.0f}', 'profit_margin_pct': '{:.1f}%',
                    'conversion_pct':    '{:.1f}%',  'total_attended': '{:,}',
                    'total_seat_booked': '{:,}'
                }).background_gradient(subset=['surplus_deficit'], cmap='RdYlGn'),
                use_container_width=True, height=380, hide_index=True
            )
        else:
            st.info("No location data available.")

    # ────────────────────────────────────────────────
    # TAB 3 — TRAINER
    # ────────────────────────────────────────────────
    with tab_trainer:
        st.markdown('<div class="section-header">👨‍🏫 Trainer Performance Analysis</div>', unsafe_allow_html=True)
        if not train_summary.empty:
            show_chart(chart_trainer_performance(train_summary), 'trainer_chart')
            st.dataframe(
                train_summary.style.format({
                    'actual_revenue':    '₹{:,.0f}', 'actual_expenses': '₹{:,.0f}',
                    'surplus_deficit':   '₹{:,.0f}', 'avg_rev_per_seminar': '₹{:,.0f}',
                    'conversion_pct':    '{:.1f}%',  'total_attended': '{:,}'
                }).background_gradient(subset=['surplus_deficit'], cmap='RdYlGn'),
                use_container_width=True, height=380, hide_index=True
            )
        else:
            st.info("No trainer data available.")

    # ────────────────────────────────────────────────
    # TAB 4 — INSIGHTS
    # ────────────────────────────────────────────────
    with tab_insights:
        st.markdown('<div class="section-header">💡 Automated Insights & Recommendations</div>', unsafe_allow_html=True)

        if not with_exp.empty and 'surplus_deficit' in with_exp.columns:
            best  = with_exp.loc[with_exp['surplus_deficit'].idxmax()]
            worst = with_exp.loc[with_exp['surplus_deficit'].idxmin()]
            loss_count   = int((with_exp['surplus_deficit'] < 0).sum())
            profit_rate  = kpi.get('profitable_seminars', 0) / kpi.get('total_seminars', 1) * 100

            ic1, ic2 = st.columns(2)
            with ic1:
                st.success(f"**🏆 Best ROI:** {best.get('location','N/A')} — Surplus ₹{int(best.get('surplus_deficit',0)):,}")
                if 'total_attended' in filtered.columns and not filtered.empty:
                    ta = filtered.loc[filtered['total_attended'].idxmax()]
                    st.info(f"**👥 Best Attendance:** {ta.get('location','N/A')} — {int(ta['total_attended']):,} attendees")
                if not train_summary.empty:
                    st.info(f"**🌟 Top Trainer:** {train_summary.iloc[0]['trainer']} — {fmt_currency(train_summary.iloc[0]['actual_revenue'])}")
                if profit_rate >= 75:
                    st.success(f"**✅ Profitability Rate: {profit_rate:.0f}%** — Excellent!")
                elif profit_rate >= 50:
                    st.warning(f"**⚠️ Profitability Rate: {profit_rate:.0f}%** — Needs improvement.")
                else:
                    st.error(f"**❌ Profitability Rate: {profit_rate:.0f}%** — Urgent review needed.")

            with ic2:
                st.error(f"**⚠️ Worst ROI:** {worst.get('location','N/A')} — Deficit ₹{abs(int(worst.get('surplus_deficit',0))):,}")
                st.warning(f"**📉 Loss-making:** {loss_count} of {len(with_exp)} seminars ran at a loss")
                exp_ratio = kpi.get('expense_ratio', 0)
                if exp_ratio > 60:
                    st.error(f"**🔴 Expense Ratio: {exp_ratio:.1f}%** — Well above 50% target.")
                elif exp_ratio > 50:
                    st.warning(f"**🟡 Expense Ratio: {exp_ratio:.1f}%** — Slightly above target.")
                else:
                    st.success(f"**🟢 Expense Ratio: {exp_ratio:.1f}%** — Within healthy range.")

            st.markdown("---")
            st.markdown("#### 📌 Action Recommendations")

            recs = []
            loss_rate = loss_count / len(with_exp) * 100 if len(with_exp) > 0 else 0
            conv = kpi.get('conversion_rate', 0)

            if loss_rate > 30:
                recs.append("🔴 **High Priority** — Over 30% of seminars are unprofitable. Immediately review cost structures and minimum revenue thresholds per venue.")
            elif loss_rate > 10:
                recs.append("🟡 **Medium Priority** — Some locations running at a loss. Conduct a cost-benefit analysis per location.")

            if conv < 20:
                recs.append(f"🔴 **Low Conversion ({conv:.1f}%)** — Less than 1 in 5 attendees is booking. Review post-seminar follow-up, pricing, and offer quality.")
            elif conv < 40:
                recs.append(f"🟡 **Average Conversion ({conv:.1f}%)** — There is room to grow. A/B test follow-up messaging and offer framing.")
            else:
                recs.append(f"🟢 **Strong Conversion ({conv:.1f}%)** — Replicate this sales process in lower-converting locations.")

            if not loc_summary.empty and len(loc_summary) > 1:
                top_loc = loc_summary.iloc[0]['location']
                recs.append(f"📍 **Expand in {top_loc}** — Your highest-revenue location. Consider adding more dates and batch times here.")

            if not train_summary.empty and len(train_summary) > 1:
                top_t   = train_summary.iloc[0]['trainer']
                low_t   = train_summary.iloc[-1]['trainer']
                recs.append(f"🌟 **Deploy {top_t} more** — Top trainer by revenue. Prioritize for high-potential cities.")
                recs.append(f"📚 **Support {low_t}** — Lowest revenue trainer. Pair with top trainer or provide sales/presentation coaching.")

            tgt = kpi.get('target_achievement', 0)
            if tgt < 70:
                recs.append(f"🎯 **Target Achievement at {tgt:.1f}%** — Review targeting strategy, marketing reach, and venue selection criteria.")
            elif tgt >= 100:
                recs.append(f"🎯 **Exceeding Target at {tgt:.1f}%** — Consider increasing targets and capacity for the next cycle.")

            pm = kpi.get('profit_margin', 0)
            if pm < 10:
                recs.append(f"💰 **Thin Margin ({pm:.1f}%)** — Explore bulk vendor discounts, reduce venue costs, or increase ticket prices marginally.")

            for rec in recs:
                st.markdown(f"> {rec}")

            # Performance Summary Table
            st.markdown("---")
            st.markdown("#### 📊 Quick Performance Summary")
            summary_data = {
                'Metric': ['Total Seminars', 'Profitable', 'Loss-making', 'Conversion %',
                           'Profit Margin %', 'Target Achievement %', 'Expense Ratio %'],
                'Value':  [kpi.get('total_seminars', 0), kpi.get('profitable_seminars', 0),
                           kpi.get('loss_seminars', 0), f"{conv:.1f}%",
                           f"{pm:.1f}%", f"{tgt:.1f}%", f"{exp_ratio:.1f}%"],
                'Status': [
                    '—',
                    '🟢' if profit_rate >= 75 else '🟡' if profit_rate >= 50 else '🔴',
                    '🔴' if loss_count > 0 else '🟢',
                    '🟢' if conv >= 40 else '🟡' if conv >= 20 else '🔴',
                    '🟢' if pm >= 20 else '🟡' if pm >= 10 else '🔴',
                    '🟢' if tgt >= 90 else '🟡' if tgt >= 70 else '🔴',
                    '🟢' if exp_ratio <= 50 else '🟡' if exp_ratio <= 60 else '🔴',
                ]
            }
            st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)
        else:
            st.info("Not enough data to generate insights.")

    # ────────────────────────────────────────────────
    # TAB 5 — DATA TABLES
    # ────────────────────────────────────────────────
    with tab_tables:
        st.markdown('<div class="section-header">📋 Seminar Performance Data</div>', unsafe_allow_html=True)
        dcols = [c for c in ['sr_no', 'location', 'trainer', 'seminar_date', 'targeted',
                              'total_attended', 'total_seat_booked', 'actual_expenses',
                              'actual_revenue', 'surplus_deficit', 'er_to_ae'] if c in filtered.columns]
        st.dataframe(
            filtered[dcols].reset_index(drop=True),
            use_container_width=True, height=400,
            column_config={
                "actual_expenses":  st.column_config.NumberColumn("Expenses",        format="₹%d"),
                "actual_revenue":   st.column_config.NumberColumn("Revenue",         format="₹%d"),
                "surplus_deficit":  st.column_config.NumberColumn("Surplus/Deficit", format="₹%d"),
                "seminar_date":     st.column_config.DateColumn("Date",              format="DD/MM/YYYY")
            }
        )

        st.markdown("---")
        st.markdown('<div class="section-header">👤 Attendee Details</div>', unsafe_allow_html=True)
        att_disp = [c for c in ['student_name', 'phone', 'email', 'service_name', 'batch_date',
                                 'payment_received', 'total_gst', 'status', 'total_amount', 'total_due']
                    if c in attendee_df.columns]
        if 'status' in attendee_df.columns:
            sf = st.multiselect("Filter by Status", attendee_df['status'].dropna().unique().tolist(), key='att_status')
            disp_att = attendee_df[attendee_df['status'].isin(sf)] if sf else attendee_df
        else:
            disp_att = attendee_df
        st.dataframe(
            disp_att[att_disp].head(500).reset_index(drop=True),
            use_container_width=True, height=400,
            column_config={
                "payment_received": st.column_config.NumberColumn("Payment",   format="₹%d"),
                "total_amount":     st.column_config.NumberColumn("Total Amt", format="₹%d"),
                "total_due":        st.column_config.NumberColumn("Due",       format="₹%d")
            }
        )
        st.caption(f"Showing {min(500, len(disp_att))} of {len(disp_att)} records")

    # ────────────────────────────────────────────────
    # TAB 6 — DOWNLOAD REPORTS
    # ────────────────────────────────────────────────
    with tab_reports:
        st.markdown('<div class="section-header">📥 Download Professional Reports</div>', unsafe_allow_html=True)

        st.markdown("""
| # | Sheet | Contents |
|---|---|---|
| 1 | 📊 Executive Summary | 12 KPIs, Top 5 & Bottom 5 locations |
| 2 | 💰 Financial Report | Full revenue / expense / surplus per seminar with totals |
| 3 | 🎯 Attendance & Conversion | Targeted → Attended → Booked with calculated conversion % |
| 4 | 👨‍🏫 Trainer Performance | Revenue, ROI, avg rev/seminar, conversion per trainer |
| 5 | 📍 Location Summary | Location-wise aggregated financials, margins & conversion |
| 6 | 👤 Attendee Details | Complete student data |
| 7 | 📚 Course Revenue | Course-wise enrollments, revenue & avg per student |
        """)

        st.markdown("")
        rc1, rc2, rc3 = st.columns([1, 2, 1])
        with rc2:
            if st.button("🔄 Generate Full Excel Report", use_container_width=True, type="primary"):
                with st.spinner("Building report — please wait…"):
                    excel_buf = generate_excel_report(
                        filtered, attendee_df, train_summary, loc_summary, course_sum, kpi
                    )
                st.success("✅ Report ready!")
                st.download_button(
                    label="📥 Download Report (.xlsx)",
                    data=excel_buf,
                    file_name=f"Seminar_Analytics_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        st.markdown("---")
        st.markdown("#### 📊 Quick CSV Downloads")
        qc1, qc2, qc3, qc4 = st.columns(4)
        with qc1:
            if not filtered.empty:
                st.download_button("📋 Seminar Data",
                                   filtered.to_csv(index=False).encode(),
                                   "seminar_data.csv", "text/csv", use_container_width=True)
        with qc2:
            if not train_summary.empty:
                st.download_button("👨‍🏫 Trainer Report",
                                   train_summary.to_csv(index=False).encode(),
                                   "trainer_report.csv", "text/csv", use_container_width=True)
        with qc3:
            if not loc_summary.empty:
                st.download_button("📍 Location Report",
                                   loc_summary.to_csv(index=False).encode(),
                                   "location_report.csv", "text/csv", use_container_width=True)
        with qc4:
            if not course_sum.empty:
                st.download_button("📚 Course Revenue",
                                   course_sum.to_csv(index=False).encode(),
                                   "course_revenue.csv", "text/csv", use_container_width=True)

    # ── Footer ──
    st.markdown("---")
    st.caption(f"**Seminar Analytics Pro** | Last refresh: {st.session_state.get('last_refresh','N/A')} | Data processed securely in-memory")

if __name__ == "__main__":
    main()
